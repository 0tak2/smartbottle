import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import os
import datetime

global app_instance_path
global exports_path

def registerInstancePathForExport(path):
    global app_instance_path
    app_instance_path = path

    global exports_path
    exports_path = os.path.join(app_instance_path, '../bottleSrv/static/exports')

    # 내보내기 패스 없으면 생성
    try:
        os.makedirs(exports_path)
    except OSError:
        pass

def tdsDataToExcel(tdsData):
    time = datetime.datetime.now()
    time_str = time.strftime('%Y-%m-%d_%H%M%S')
    fileName = 'TDS_' + time_str + '.xlsx'

    df = makeDf(tdsData, ['ID', '입력시간', 'TDS 수치(mg/L)'])
    return dfToExcel(df, fileName, 'TDS 데이터 내보내기')

def hydrationDataToExcel(hydrData):
    time = datetime.datetime.now()
    time_str = time.strftime('%Y-%m-%d_%H%M%S')
    fileName = 'HYDR_' + time_str + '.xlsx'

    df = makeDf(hydrData, ['ID', '입력시간', '부피 변화(ml)', '현재 부피(ml)'])
    return dfToExcel(df, fileName, 'Hydration 데이터 내보내기')

def makeDf(TwoDList, header):
    df = pd.DataFrame(TwoDList, columns = header)
    return df

def dfToExcel(df, fileName, sheet_name):
    global exports_path
    filePath = os.path.join(exports_path, fileName)

    df.to_excel(filePath, sheet_name=sheet_name, index=False)

    # 스타일링을 위해 다시 파일을 연다
    wb = load_workbook(filePath)
    ws = wb.active

    # 각 열에 최적의 너비를 설정한다
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20

    # 모든 셀에 자동 줄바꿈, 세로 정렬을 설정한다
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

   # 헤더에 가운데 정렬을 설정한다
    for cell in ws[1]:
        cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")

    # A열에 가운데 정렬을 설정한다
    for row in ws.iter_rows():
        row[0].alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")

    wb.save(filePath)

    return '/static/exports/' + fileName